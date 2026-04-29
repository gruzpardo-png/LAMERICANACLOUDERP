import os
import csv
import io
import hmac
import hashlib
import secrets
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from sqlalchemy import create_engine, Column, Integer, String, Float, Boolean, DateTime, ForeignKey, func
from sqlalchemy.orm import declarative_base, sessionmaker, Session, relationship
from starlette.middleware.sessions import SessionMiddleware
from openpyxl import load_workbook

APP_NAME = "LA ERP Cloud"
FAMILIES = ["vestuario", "hogar", "zapatillas", "bolsos"]
TERMINALS = ["T1", "T2", "T3", "T4"]

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./lamericana.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql+psycopg2://", 1)
elif DATABASE_URL.startswith("postgresql://") and "+" not in DATABASE_URL.split("://", 1)[0]:
    DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+psycopg2://", 1)

SECRET_KEY = os.getenv("SECRET_KEY", secrets.token_urlsafe(32))
AGENT_TOKEN = os.getenv("AGENT_TOKEN", "cambiar-token-agente")
APP_TIMEZONE = os.getenv("APP_TIMEZONE", "America/Santiago")

engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {},
    pool_pre_ping=True,
)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
Base = declarative_base()
app = FastAPI(title=APP_NAME)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, same_site="lax")


def now():
    try:
        return datetime.now(ZoneInfo(APP_TIMEZONE)).replace(tzinfo=None)
    except Exception:
        return datetime.now()


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    full_name = Column(String(120), nullable=False)
    code = Column(String(10), nullable=False)
    role = Column(String(20), default="operator")
    active = Column(Boolean, default=True)
    labels = relationship("Label", back_populates="user")


class PriceItem(Base):
    __tablename__ = "price_items"
    id = Column(Integer, primary_key=True)
    family = Column(String(30), index=True, nullable=False)
    product_code = Column(String(80), index=True, nullable=False)
    description = Column(String(255), nullable=False)
    price_gross = Column(Float, index=True, nullable=False)
    active = Column(Boolean, default=True)


class TerminalWeight(Base):
    __tablename__ = "terminal_weights"
    id = Column(Integer, primary_key=True)
    terminal = Column(String(20), unique=True, index=True, nullable=False)
    weight_kg = Column(Float, default=0.0)
    source = Column(String(50), default="manual")
    updated_at = Column(DateTime, default=now)


class Label(Base):
    __tablename__ = "labels"
    id = Column(Integer, primary_key=True)
    created_at = Column(DateTime, default=now, index=True)
    user_id = Column(Integer, ForeignKey("users.id"))
    username = Column(String(50), nullable=False)
    user_code = Column(String(10), nullable=False)
    terminal = Column(String(20), index=True, nullable=False)
    family = Column(String(30), index=True, nullable=False)
    method = Column(String(20), nullable=False)
    weight_kg = Column(Float, nullable=False)
    value_per_kg = Column(Float)
    target_price = Column(Float)
    final_price = Column(Float, nullable=False)
    product_code = Column(String(80), nullable=False)
    description = Column(String(255), nullable=False)
    idx_value = Column(Float)
    in_origin = Column(String(255))
    print_status = Column(String(30), default="no_solicitada")
    user = relationship("User", back_populates="labels")


class PriceImport(Base):
    __tablename__ = "price_imports"
    id = Column(Integer, primary_key=True)
    created_at = Column(DateTime, default=now, index=True)
    username = Column(String(50), nullable=False)
    file_name = Column(String(255), nullable=False)
    total_rows = Column(Integer, default=0)
    inserted_rows = Column(Integer, default=0)
    skipped_rows = Column(Integer, default=0)
    status = Column(String(30), default="ok")
    detail = Column(String(2000), default="")


def db_session():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def hash_password(password):
    salt = secrets.token_hex(16)
    iterations = 260000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), iterations).hex()
    return f"pbkdf2_sha256${iterations}${salt}${digest}"


def verify_password(password, hashed):
    try:
        scheme, iterations, salt, digest = hashed.split("$", 3)
        if scheme != "pbkdf2_sha256":
            return False
        candidate = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), int(iterations)).hex()
        return hmac.compare_digest(candidate, digest)
    except Exception:
        return False


def money(value):
    try:
        return f"${int(round(float(value))):,}".replace(",", ".")
    except Exception:
        return str(value)


def to_float(value, default=None):
    if value is None:
        return default
    text = str(value).strip().replace("$", "").replace(".", "").replace(",", ".")
    if not text:
        return default
    try:
        return float(text)
    except Exception:
        return default


def current_user(request: Request, db: Session):
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    user = db.get(User, user_id)
    if not user or not user.active:
        request.session.clear()
        return None
    return user


def terminal(request: Request):
    return request.session.get("terminal", "T1")


def page(title, body, request, user=None):
    nav = ""
    session = ""
    if user:
        nav = """
        <a href='/dashboard'>Dashboard</a>
        <a href='/etiquetas'>Etiquetas</a>
        <a href='/historial'>Historial</a>
        <a href='/precios'>Precios</a>
        <a href='/logout'>Salir</a>
        """
        session = f"<div class='session'>Usuario: <b>{user.code}</b> {user.full_name} · Terminal: <b>{terminal(request)}</b></div>"
    html = f"""
    <!doctype html><html lang='es'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width, initial-scale=1'>
    <title>{title}</title>
    <style>
    body{{margin:0;font-family:Arial;background:#f5f7fb;color:#1f2937}}a{{color:#8b1d3d;font-weight:700;text-decoration:none;margin-right:14px}}
    header{{background:#111827;color:white;padding:14px 22px;display:flex;justify-content:space-between;gap:16px;flex-wrap:wrap}}header a{{color:white}}
    .session{{background:#fff7ed;border-bottom:1px solid #fed7aa;padding:8px 22px}}.container{{width:min(1200px,calc(100% - 32px));margin:22px auto 60px}}
    .panel,.card,.login{{background:white;border:1px solid #e5e7eb;border-radius:16px;padding:18px;margin-bottom:16px;box-shadow:0 6px 18px rgba(0,0,0,.04)}}
    .cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px}}.card span{{color:#6b7280;display:block}}.card strong{{font-size:30px;color:#8b1d3d}}
    input,select{{width:100%;padding:10px;border:1px solid #d1d5db;border-radius:10px;font-size:15px}}label{{display:grid;gap:6px;font-weight:700;margin-bottom:12px}}
    button,.btn{{background:#8b1d3d;color:white;border:0;border-radius:10px;padding:10px 14px;font-weight:800;cursor:pointer;display:inline-block}}
    table{{width:100%;border-collapse:collapse;background:white}}th,td{{border-bottom:1px solid #e5e7eb;padding:9px;text-align:left}}th{{background:#f9fafb}}
    .grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}.alert{{background:#fee2e2;color:#991b1b;border:1px solid #fecaca;padding:10px;border-radius:10px}}
    .price{{font-size:46px;font-weight:900;text-align:center}}.labelbox{{border:2px dashed #111827;border-radius:12px;padding:14px;max-width:380px;background:white}}
    @media(max-width:850px){{.grid{{grid-template-columns:1fr}}}}
    </style></head><body><header><b>LA ERP Cloud · La Americana</b><nav>{nav}</nav></header>{session}<main class='container'>{body}</main></body></html>
    """
    return HTMLResponse(html)


def require_login(request, db):
    user = current_user(request, db)
    if not user:
        return None, RedirectResponse("/login", status_code=303)
    return user, None


def find_product(db, family, amount, upwards=False):
    q = db.query(PriceItem).filter(PriceItem.family == family, PriceItem.active == True)
    if upwards:
        product = q.filter(PriceItem.price_gross >= amount).order_by(PriceItem.price_gross.asc()).first()
        return product or q.order_by(PriceItem.price_gross.desc()).first()
    items = q.all()
    return min(items, key=lambda p: abs(p.price_gross - amount)) if items else None


def normalize_family(name):
    s = str(name or "").strip().lower()
    aliases = {
        "vestuario": "vestuario",
        "ropa": "vestuario",
        "hogar": "hogar",
        "casa": "hogar",
        "zapatillas": "zapatillas",
        "zapatos": "zapatillas",
        "calzado": "zapatillas",
        "bolsos": "bolsos",
        "bolso": "bolsos",
        "carteras": "bolsos",
    }
    return aliases.get(s)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_product_code(value):
    text = clean_text(value)
    if text.endswith(".0") and text.replace(".0", "").isdigit():
        text = text[:-2]
    return text


def detect_price_columns(sample_rows):
    headers = [clean_text(v).lower() for v in sample_rows[0]] if sample_rows else []
    idx_code = idx_desc = idx_price = None
    for i, h in enumerate(headers):
        compact = h.replace(" ", "").replace("_", "")
        if idx_code is None and ("codigo" in compact or "código" in compact or compact in ("cod", "sku")):
            idx_code = i
        if idx_desc is None and ("descripcion" in compact or "descripción" in compact or compact in ("desc", "producto", "nombre")):
            idx_desc = i
        if idx_price is None and ("precio" in compact or "valor" in compact or "venta" in compact):
            idx_price = i
    if idx_code is not None and idx_desc is not None and idx_price is not None:
        return idx_code, idx_desc, idx_price, 2
    return 0, 1, 2, 1


def parse_prices_excel(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    imported = []
    skipped = []
    sheet_summary = {}
    found_families = set()

    for ws in wb.worksheets:
        family = normalize_family(ws.title)
        if not family:
            continue

        found_families.add(family)
        first_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            first_rows.append(row)
            if i >= 5:
                break

        idx_code, idx_desc, idx_price, start_row = detect_price_columns(first_rows)
        inserted_sheet = 0
        skipped_sheet = 0

        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_number < start_row:
                continue
            max_idx = max(idx_code, idx_desc, idx_price)
            if len(row) <= max_idx:
                skipped_sheet += 1
                continue

            code = normalize_product_code(row[idx_code])
            description = clean_text(row[idx_desc])
            price = to_float(row[idx_price])

            if not code and not description and price is None:
                continue

            if not code or not description or price is None or price <= 0:
                skipped_sheet += 1
                if len(skipped) < 30:
                    skipped.append(f"{ws.title} fila {row_number}: código/descripción/precio inválido")
                continue

            imported.append({
                "family": family,
                "product_code": code[:80],
                "description": description[:255],
                "price_gross": float(price),
            })
            inserted_sheet += 1

        sheet_summary[family] = inserted_sheet
        if skipped_sheet:
            sheet_summary[f"{family}_omitidas"] = skipped_sheet

    missing = [f for f in FAMILIES if f not in found_families]
    return imported, skipped, sheet_summary, missing


def price_family_stats(db):
    rows = []
    for f in FAMILIES:
        total = db.query(PriceItem).filter(PriceItem.family == f, PriceItem.active == True).count()
        min_price = db.query(func.min(PriceItem.price_gross)).filter(PriceItem.family == f, PriceItem.active == True).scalar()
        max_price = db.query(func.max(PriceItem.price_gross)).filter(PriceItem.family == f, PriceItem.active == True).scalar()
        rows.append({"family": f, "total": total, "min": min_price or 0, "max": max_price or 0})
    return rows


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if not db.query(User).filter_by(username="gustavo").first():
            db.add(User(username="gustavo", password_hash=hash_password("1176"), full_name="Gustavo", code="GS", role="admin"))
        for t in TERMINALS:
            if not db.query(TerminalWeight).filter_by(terminal=t).first():
                db.add(TerminalWeight(terminal=t, weight_kg=0.0, source="inicial"))
        if db.query(PriceItem).count() == 0:
            base_prices = [("vestuario","V1990","PRENDA VESTUARIO",1990),("vestuario","V2990","PRENDA VESTUARIO",2990),("vestuario","V3990","PRENDA VESTUARIO",3990),("hogar","H2990","ARTICULO HOGAR",2990),("hogar","H4990","ARTICULO HOGAR",4990),("zapatillas","Z9990","ZAPATILLAS",9990),("zapatillas","Z14990","ZAPATILLAS",14990),("bolsos","B4990","BOLSO",4990)]
            for f,c,d,p in base_prices:
                db.add(PriceItem(family=f, product_code=c, description=d, price_gross=float(p), active=True))
        db.commit()
    finally:
        db.close()


@app.get("/")
def root(request: Request, db: Session = Depends(db_session)):
    return RedirectResponse("/dashboard" if current_user(request, db) else "/login", status_code=303)


@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request):
    opts = "".join(f"<option value='{t}'>{t}</option>" for t in TERMINALS)
    body = f"""
    <section class='login'><h1>Ingreso LA ERP Cloud</h1><p>Usuario inicial: <b>gustavo</b> · clave: <b>1176</b></p>
    <form method='post'><label>Usuario<input name='username' required autofocus></label><label>Clave<input name='password' type='password' required></label><label>Terminal<select name='term'>{opts}</select></label><button>Ingresar</button></form></section>
    """
    return page("Login", body, request)


@app.post("/login")
def login_post(request: Request, username: str = Form(...), password: str = Form(...), term: str = Form("T1"), db: Session = Depends(db_session)):
    user = db.query(User).filter_by(username=username.strip().lower(), active=True).first()
    if not user or not verify_password(password.strip(), user.password_hash):
        return page("Error", "<div class='alert'>Usuario o clave incorrecta.</div><a href='/login'>Volver</a>", request)
    request.session["user_id"] = user.id
    request.session["terminal"] = term if term in TERMINALS else "T1"
    return RedirectResponse("/dashboard", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir: return redir
    start = datetime.combine(now().date(), datetime.min.time())
    end = start + timedelta(days=1)
    q = db.query(Label).filter(Label.created_at >= start, Label.created_at < end)
    total = q.count()
    kg = q.with_entities(func.coalesce(func.sum(Label.weight_kg), 0)).scalar() or 0
    value = q.with_entities(func.coalesce(func.sum(Label.final_price), 0)).scalar() or 0
    pending = db.query(Label).filter_by(print_status="pendiente").count()
    latest = db.query(Label).order_by(Label.id.desc()).limit(15).all()
    rows = "".join(f"<tr><td>{l.created_at.strftime('%H:%M:%S')}</td><td>{l.user_code}</td><td>{l.terminal}</td><td>{l.family}</td><td>{l.weight_kg:.3f}</td><td>{money(l.final_price)}</td><td>{l.print_status}</td></tr>" for l in latest) or "<tr><td colspan='7'>Sin datos</td></tr>"
    body = f"""<h1>Dashboard</h1><section class='cards'><div class='card'><span>Etiquetas hoy</span><strong>{total}</strong></div><div class='card'><span>Kilos hoy</span><strong>{float(kg):.3f}</strong></div><div class='card'><span>Valor proyectado</span><strong>{money(value)}</strong></div><div class='card'><span>Pendientes impresión</span><strong>{pending}</strong></div></section><section class='panel'><h2>Últimas etiquetas</h2><table><tr><th>Hora</th><th>Usuario</th><th>Terminal</th><th>Familia</th><th>Peso</th><th>Precio</th><th>Estado</th></tr>{rows}</table></section>"""
    return page("Dashboard", body, request, user)


@app.get("/etiquetas", response_class=HTMLResponse)
def etiquetas(request: Request, label_id: int = 0, db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir: return redir
    t = terminal(request)
    tw = db.query(TerminalWeight).filter_by(terminal=t).first()
    label = db.get(Label, label_id) if label_id else None
    fopts = "".join(f"<option value='{f}'>{f.capitalize()}</option>" for f in FAMILIES)
    result = "<p>Calcula una etiqueta para ver resultado.</p>"
    if label:
        result = f"<div class='labelbox'><div class='price'>{money(label.final_price)}</div><hr><b>Cod:</b> {label.product_code}<br><b>Peso:</b> {label.weight_kg:.3f} kg<br><b>IDX:</b> {(label.idx_value or 0):.2f} {label.user_code} {label.terminal}<br><b>IN:</b> {label.in_origin or '--'}<br><b>DS:</b> {label.description}<br><b>Estado:</b> {label.print_status}</div>"
    body = f"""
    <h1>Etiquetado</h1><section class='grid'><div class='panel'><h2>Nueva etiqueta</h2><div class='card'><span>Peso actual {t}</span><strong>{(tw.weight_kg if tw else 0):.3f} kg</strong></div>
    <form method='post' action='/etiquetas/calcular'><label>Familia<select name='family'>{fopts}</select></label><label>Método<select name='method'><option value='kilo'>Por kilo</option><option value='objetivo'>Precio objetivo</option></select></label><label>Fuente peso<select name='weight_source'><option value='agent'>Balanza/agente</option><option value='manual'>Manual</option></select></label><label>Peso manual kg<input name='manual_weight_kg' placeholder='0.352'></label><label>Valor kilo<input name='value_per_kg' placeholder='12000'></label><label>Precio objetivo<input name='target_price' placeholder='4990'></label><label>IN / Procedencia<input name='in_origin' placeholder='Fardo / proveedor / lote'></label><label><input type='checkbox' name='print_requested' value='1' style='width:auto'> Solicitar impresión local</label><button>Calcular / Guardar</button></form></div><div class='panel'><h2>Resultado</h2>{result}</div></section>
    """
    return page("Etiquetas", body, request, user)


@app.post("/etiquetas/calcular")
def calcular(request: Request, family: str = Form(...), method: str = Form(...), weight_source: str = Form("agent"), manual_weight_kg: str = Form(""), value_per_kg: str = Form(""), target_price: str = Form(""), in_origin: str = Form(""), print_requested: str = Form(None), db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir: return redir
    t = terminal(request)
    family = family.lower().strip()
    tw = db.query(TerminalWeight).filter_by(terminal=t).first()
    weight = to_float(manual_weight_kg) if weight_source == "manual" else (tw.weight_kg if tw else 0)
    if not weight or weight <= 0:
        return page("Error", "<div class='alert'>Peso inválido. Ingresa peso manual o conecta el agente local.</div><a href='/etiquetas'>Volver</a>", request, user)
    if method == "kilo":
        value_kg = to_float(value_per_kg)
        if not value_kg:
            return page("Error", "<div class='alert'>Valor kilo inválido.</div><a href='/etiquetas'>Volver</a>", request, user)
        product = find_product(db, family, value_kg * weight, upwards=False)
        target = None
        idx = value_kg / 1000
    else:
        target = to_float(target_price)
        if not target:
            return page("Error", "<div class='alert'>Precio objetivo inválido.</div><a href='/etiquetas'>Volver</a>", request, user)
        product = find_product(db, family, target, upwards=True)
        value_kg = product.price_gross / weight if product else None
        idx = value_kg / 1000 if value_kg else None
    if not product:
        return page("Error", "<div class='alert'>No hay precios cargados para esa familia.</div><a href='/precios'>Ir a precios</a>", request, user)
    label = Label(user_id=user.id, username=user.username, user_code=user.code, terminal=t, family=family, method=method, weight_kg=float(weight), value_per_kg=value_kg, target_price=target, final_price=product.price_gross, product_code=product.product_code, description=product.description, idx_value=idx, in_origin=in_origin.strip() or None, print_status="pendiente" if print_requested else "no_solicitada")
    db.add(label); db.commit(); db.refresh(label)
    return RedirectResponse(f"/etiquetas?label_id={label.id}", status_code=303)


@app.get("/historial", response_class=HTMLResponse)
def historial(request: Request, db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir: return redir
    labels = db.query(Label).order_by(Label.id.desc()).limit(300).all()
    rows = "".join(f"<tr><td>{l.id}</td><td>{l.created_at.strftime('%d-%m-%Y %H:%M:%S')}</td><td>{l.user_code}</td><td>{l.terminal}</td><td>{l.family}</td><td>{l.weight_kg:.3f}</td><td>{money(l.final_price)}</td><td>{l.product_code}</td><td>{l.description}</td><td>{l.print_status}</td></tr>" for l in labels) or "<tr><td colspan='10'>Sin etiquetas</td></tr>"
    body = f"<h1>Historial</h1><a class='btn' href='/export/labels.csv'>Exportar CSV</a><section class='panel'><table><tr><th>ID</th><th>Fecha</th><th>Usuario</th><th>Terminal</th><th>Familia</th><th>Peso</th><th>Precio</th><th>Código</th><th>Descripción</th><th>Estado</th></tr>{rows}</table></section>"
    return page("Historial", body, request, user)


@app.get("/export/labels.csv")
def export_csv(request: Request, db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir: return redir
    out = io.StringIO(); writer = csv.writer(out, delimiter=";")
    writer.writerow(["id","fecha","usuario","terminal","familia","metodo","peso_kg","precio","codigo","descripcion","estado"])
    for l in db.query(Label).order_by(Label.id.desc()).all():
        writer.writerow([l.id,l.created_at.strftime("%Y-%m-%d %H:%M:%S"),l.user_code,l.terminal,l.family,l.method,l.weight_kg,l.final_price,l.product_code,l.description,l.print_status])
    out.seek(0)
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=historial_etiquetas.csv"})


@app.get("/precios", response_class=HTMLResponse)
def precios(request: Request, family: str = "", q: str = "", db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir:
        return redir

    family_filter = normalize_family(family) if family else ""
    query = db.query(PriceItem).filter(PriceItem.active == True)
    if family_filter:
        query = query.filter(PriceItem.family == family_filter)
    q_text = q.strip()
    if q_text:
        like = f"%{q_text}%"
        query = query.filter((PriceItem.product_code.ilike(like)) | (PriceItem.description.ilike(like)))

    items = query.order_by(PriceItem.family, PriceItem.price_gross, PriceItem.product_code).limit(500).all()
    stats = price_family_stats(db)
    last_import = db.query(PriceImport).order_by(PriceImport.id.desc()).first()

    cards = "".join(
        f"<div class='card'><span>{s['family'].capitalize()}</span><strong>{s['total']}</strong><small>{money(s['min'])} a {money(s['max'])}</small></div>"
        for s in stats
    )

    family_options = "<option value=''>Todas las familias</option>" + "".join(
        f"<option value='{f}' {'selected' if f == family_filter else ''}>{f.capitalize()}</option>" for f in FAMILIES
    )

    rows = "".join(
        f"<tr><td><b>{p.family}</b></td><td>{p.product_code}</td><td>{p.description}</td><td><b>{money(p.price_gross)}</b></td></tr>"
        for p in items
    ) or "<tr><td colspan='4'>Sin precios cargados para el filtro actual.</td></tr>"

    import_info = ""
    if last_import:
        import_info = (
            f"<div class='card'><span>Última actualización</span>"
            f"<strong>{last_import.inserted_rows}</strong>"
            f"<small>{last_import.created_at.strftime('%d-%m-%Y %H:%M')} · {last_import.file_name} · omitidas: {last_import.skipped_rows}</small></div>"
        )

    admin_upload = ""
    if user.role == "admin":
        admin_upload = """
        <section class='panel'>
            <h2>Actualizar precios desde Excel</h2>
            <p>Sube la planilla base con hojas <b>vestuario</b>, <b>hogar</b>, <b>zapatillas</b> y <b>bolsos</b>. El sistema lee las primeras 3 columnas: <b>Código</b>, <b>Descripción</b> y <b>Precio Venta Bruto</b>. También acepta encabezados si existen.</p>
            <form method='post' action='/precios/importar' enctype='multipart/form-data'>
                <label>Archivo Excel .xlsx
                    <input type='file' name='file' accept='.xlsx' required>
                </label>
                <label><input type='checkbox' name='confirm_replace' value='1' style='width:auto' required> Confirmo reemplazar completamente la tabla de precios actual</label>
                <button>Importar y actualizar precios</button>
            </form>
        </section>
        """

    body = f"""
    <h1>Precios base</h1>
    <section class='cards'>{cards}{import_info}</section>
    {admin_upload}
    <section class='panel'>
        <h2>Consulta de precios cargados</h2>
        <form method='get' style='display:grid;grid-template-columns:220px 1fr 160px;gap:12px;align-items:end'>
            <label>Familia<select name='family'>{family_options}</select></label>
            <label>Buscar código/descripción<input name='q' value='{q_text}' placeholder='Ej: ITA07/VES o 76750'></label>
            <button>Filtrar</button>
        </form>
        <p><a class='btn' href='/export/precios.csv'>Exportar precios CSV</a></p>
        <table><tr><th>Familia</th><th>Código</th><th>Descripción</th><th>Precio</th></tr>{rows}</table>
        <p style='color:#6b7280'>Mostrando máximo 500 filas. Usa el filtro para buscar códigos específicos.</p>
    </section>
    """
    return page("Precios", body, request, user)


@app.post("/precios/importar")
async def importar_precios(request: Request, file: UploadFile = File(...), confirm_replace: str = Form(None), db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir:
        return redir
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden importar precios")
    if confirm_replace != "1":
        return page("Error", "<div class='alert'>Debes confirmar el reemplazo completo de precios.</div><a href='/precios'>Volver</a>", request, user)
    if not file.filename.lower().endswith(".xlsx"):
        return page("Error", "<div class='alert'>Debes subir un archivo Excel .xlsx.</div><a href='/precios'>Volver</a>", request, user)

    try:
        content = await file.read()
        items, skipped, sheet_summary, missing = parse_prices_excel(content)
        if not items:
            return page("Error", "<div class='alert'>No se encontraron precios válidos en el Excel. Revisa hojas y columnas.</div><a href='/precios'>Volver</a>", request, user)

        db.query(PriceItem).delete()
        for item in items:
            db.add(PriceItem(**item, active=True))

        detail_parts = [f"{k}: {v}" for k, v in sheet_summary.items()]
        skipped_count = sum(v for k, v in sheet_summary.items() if k.endswith("_omitidas"))
        if missing:
            detail_parts.append("Hojas no encontradas: " + ", ".join(missing))
        if skipped:
            detail_parts.append("Omitidas: " + " | ".join(skipped[:10]))

        log = PriceImport(
            username=user.username,
            file_name=file.filename,
            total_rows=len(items) + skipped_count,
            inserted_rows=len(items),
            skipped_rows=skipped_count,
            status="ok",
            detail="; ".join(detail_parts)[:2000],
        )
        db.add(log)
        db.commit()

        resumen_familias = "".join(
            f"<tr><td>{f}</td><td>{sheet_summary.get(f, 0)}</td><td>{sheet_summary.get(f + '_omitidas', 0)}</td></tr>"
            for f in FAMILIES
        )
        body = f"""
        <h1>Precios actualizados</h1>
        <section class='panel'>
            <div class='cards'>
                <div class='card'><span>Registros importados</span><strong>{len(items)}</strong></div>
                <div class='card'><span>Filas omitidas</span><strong>{skipped_count}</strong></div>
                <div class='card'><span>Archivo</span><strong style='font-size:18px'>{file.filename}</strong></div>
            </div>
            <h2>Resumen por familia</h2>
            <table><tr><th>Familia</th><th>Importadas</th><th>Omitidas</th></tr>{resumen_familias}</table>
            <p><a class='btn' href='/precios'>Ver precios cargados</a> <a class='btn' href='/etiquetas'>Ir a etiquetar</a></p>
        </section>
        """
        return page("Precios actualizados", body, request, user)
    except Exception as exc:
        db.rollback()
        db.add(PriceImport(username=user.username, file_name=file.filename, total_rows=0, inserted_rows=0, skipped_rows=0, status="error", detail=str(exc)[:2000]))
        db.commit()
        return page("Error", f"<div class='alert'>No se pudo importar el Excel: {str(exc)}</div><a href='/precios'>Volver</a>", request, user)


@app.get("/export/precios.csv")
def export_precios_csv(request: Request, db: Session = Depends(db_session)):
    user, redir = require_login(request, db)
    if redir:
        return redir
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["familia", "codigo_producto", "descripcion", "precio_venta_bruto", "activo"])
    for p in db.query(PriceItem).order_by(PriceItem.family, PriceItem.price_gross, PriceItem.product_code).all():
        writer.writerow([p.family, p.product_code, p.description, int(round(p.price_gross)), "SI" if p.active else "NO"])
    out.seek(0)
    return StreamingResponse(iter([out.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=precios_base_lamericana.csv"})


@app.post("/api/agent/weight")
async def api_weight(request: Request, db: Session = Depends(db_session)):
    token = request.headers.get("X-Agent-Token") or request.query_params.get("token")
    if token != AGENT_TOKEN:
        raise HTTPException(status_code=403, detail="Token inválido")
    data = await request.json()
    t = str(data.get("terminal", "T1")).strip()
    weight = to_float(data.get("weight_kg"), 0.0) or 0.0
    if t not in TERMINALS:
        return JSONResponse({"ok": False, "error": "Terminal inválido"}, status_code=400)
    tw = db.query(TerminalWeight).filter_by(terminal=t).first() or TerminalWeight(terminal=t)
    db.add(tw)
    tw.weight_kg = float(weight); tw.source = "agent"; tw.updated_at = now()
    db.commit()
    return {"ok": True, "terminal": t, "weight_kg": tw.weight_kg}


@app.get("/api/agent/next-print")
def api_next_print(request: Request, terminal: str = "T1", db: Session = Depends(db_session)):
    token = request.headers.get("X-Agent-Token") or request.query_params.get("token")
    if token != AGENT_TOKEN:
        raise HTTPException(status_code=403, detail="Token inválido")
    l = db.query(Label).filter_by(terminal=terminal, print_status="pendiente").order_by(Label.id).first()
    if not l:
        return {"ok": True, "label": None}
    return {"ok": True, "label": {"id": l.id, "price_text": money(l.final_price), "price": l.final_price, "weight_kg": l.weight_kg, "product_code": l.product_code, "description": l.description, "idx_value": l.idx_value, "user_code": l.user_code, "terminal": l.terminal, "in_origin": l.in_origin}}


@app.post("/api/agent/print-result")
async def api_print_result(request: Request, db: Session = Depends(db_session)):
    token = request.headers.get("X-Agent-Token") or request.query_params.get("token")
    if token != AGENT_TOKEN:
        raise HTTPException(status_code=403, detail="Token inválido")
    data = await request.json()
    l = db.get(Label, int(data.get("label_id", 0)))
    if not l:
        return JSONResponse({"ok": False, "error": "Etiqueta no encontrada"}, status_code=404)
    l.print_status = "impresa" if data.get("ok") else "error_impresion"
    db.commit()
    return {"ok": True}


@app.get("/health")
def health():
    return {"ok": True, "app": APP_NAME, "time": now().isoformat()}
